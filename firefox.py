import os, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from PIL import ImageGrab
import win32com.client as win32

from selenium import webdriver
from selenium.webdriver import Firefox
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service

from datetime import datetime

start_time = datetime.now()

#=== CRIANDO UM PROFILE NO FIREFOX ===#

# Windows + R
# "firefox.exe -p"
# Crie um novo perfil
# Entre na pasta: "%appdata%\Roaming\Mozilla\Firefox\Profiles"
# Copie o caminho para a variavel "profile_path"

driver_path = str(Path(__file__).parent.absolute()) + r"\geckodriver.exe" #caminho do navegador
profile_path = r"C:\Users\<seu_usuario>\AppData\Roaming\Mozilla\Firefox\Profiles\<seu_profile>" #caminho do profile, geralmente localizado em "%appdata%\Roaming\Mozilla\Firefox\Profiles" *fazer login no whatsapp antes

lista_contatos = []
lista_figuras = []

#=== NOME DA PLANILHA USADA PARA RODAR O ENVIO ===#
planilha_contatos = str(Path(__file__).parent.absolute()) + r"\Exemplo.xlsx"
aba_planilha = "Aba1"

#Nesse código os contatos ficam a esquerda e o nome da imagem fica a direita, tudo na mesma planilha

#=== INICIANDO O NAVEGADOR ===#
options = Options()
#options.headless = True #pra rodar o firefox em modo oculto
options.add_argument("-profile") #adicionando a pasta profile p/ o whatsapp não ficar pedindo o QR code 
options.add_argument(profile_path)

service = Service(driver_path)

driver = Firefox(service=service, options=options)

driver.get("http://web.whatsapp.com")

wait = WebDriverWait(driver, 800)

def importar_contatos(planilha):
    wb = load_workbook(planilha)

    try:
        ws = wb[aba_planilha] #nome da aba da planilha
        coluna = 1 #coluna A
    
        for row in range(2, ws.max_row+1): #numero da linha onde começa / maximo de linhas +1 para não finalizar antes do ultimo elemento   
            if(ws.cell(row, coluna).value is None): #se valor é nulo -> break
                break
            else:
                lista_contatos.append(ws.cell(row, coluna).value) #concatena o valor encontrado na lista
    except Exception as e:
        print("Erro ao importar contatos: " + str(e))
    finally:
        wb.close() #fecha o workbook p/ win32com conseguir copiar a imagem

def importar_figura(planilha):
    wb = load_workbook(planilha)

    try:
        ws = wb[aba_planilha] #nome da aba da planilha
        coluna = 2 #coluna B
    
        for row in range(2, ws.max_row+1): #numero da linha onde começa / maximo de linhas +1 para não finalizar antes do ultimo elemento
            if(ws.cell(row, coluna).value is None): #se valor é nulo -> break
                break
            else:
                lista_figuras.append(ws.cell(row, coluna).value) #define na variavel o nome da imagem
    except Exception as e:
        print("Erro ao importar nome da imagem: " + str(e))
    finally:
        wb.close() #fecha o workbook p/ win32com conseguir copiar a imagem

def salvar_imagem(planilha):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(planilha)

    try:
        sheet = excel.Sheets(aba_planilha)

        for i, shape in enumerate(sheet.Shapes): #percorre todos os shapes da pasta selecionada
            for figura in lista_figuras: #percorre todas as figuras armazenadas na lista de figuras
                if shape.Name.startswith(figura): #se o nome da figura for igual ao elemento da lista de figuras
                    shape.Copy()
                    image = ImageGrab.grabclipboard() #copia
                    image = image.convert('RGB') #converte de RGBA p/ RGB
                    image.save(figura + ".jpg", "jpeg") #salva com a extensao .jpg
    except Exception as e:
        print("Erro ao salvar a imagem: " + str(e))
    finally:
        wb.Close(True) #fecha o workbook
        excel.Quit() #finaliza o processo do excel

def envia_imagens(contatos, imagens):
    print("Enviando mensagens...")
    time.sleep(5)
    try:
        for i in range(0, len(contatos)): #percorre do 0 até o tamanho do array dos contatos enviando para o contato a respectiva imagem na mesma linha da planilha
            print("Enviando imagem: ## " + str(imagens[i].upper()) + " ## Para: ## " + str(contatos[i]).upper() + " ##")
            #procura na pagina o elemento com o mesmo nome do contato
            x_arg = '//span[contains(@title, ' + '"' + str(contatos[i]) + '"' + ')]'
            group_title = wait.until(EC.presence_of_element_located((
                By.XPATH, x_arg)))
            group_title.click()
            #procura na pagina o elemento de anexo de arquivos
            driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
            attach = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
            #passa o caminho da imagem para ser anexada
            attach.send_keys(str(Path(__file__).parent.absolute()) + "\\" + str(imagens[i]) + ".jpg")
            time.sleep(3)
            #tecla enviar
            send = driver.find_element(By.CSS_SELECTOR, "span[data-icon='send']")
            send.click()
            time.sleep(5)
    except Exception as e:
        print("Erro ao enviar as mensagens: " + str(e))
    finally:
        time.sleep(3)
        driver.close() #fecha o driver  

#importando contatos
importar_contatos(planilha_contatos) #passando a planilha carregada antes como argumento
print("Lista de contatos importada:") 
print(lista_contatos)

#importando figuras
importar_figura(planilha_contatos) #passando a planilha carregada antes como argumento
print("Lista de figuras importadas:")
print(lista_figuras)

salvar_imagem(planilha_contatos) #passando a planilha carregada antes como argumento

envia_imagens(lista_contatos, lista_figuras) #passa a lista de contatos e a lista de figuras armazenadas nos arrays como argumentos

#excluindo as figuras geradas
for figura in lista_figuras:
    os.remove(str(figura) + ".jpg")

end_time = datetime.now()

print('---- Tempo de execução do script: {} ----'.format(end_time - start_time))