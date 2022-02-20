import os, time, shutil
from posixpath import abspath
from pathlib import Path

from openpyxl import load_workbook
from PIL import ImageGrab
import win32com.client as win32
#Excel2Img

from selenium.webdriver import Firefox
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.keys import Keys 

from datetime import datetime, timedelta
from prettytable import PrettyTable

#=============== VARIAVEIS GLOBAIS ===============#
contEnvios = 0
#Variaveis de cordenada p/ os arrays passados como argumento
Linha = 1
Coluna = 0
#Variaveis temporarias p/ armazenar os nomes dos contatos, figuras e as datas
lstContatos = []
lstFiguras = []
lstDatas = []
lstContatosErros = []
#Tabela tabAtrasos com os atrasos
tabAtrasos = PrettyTable(["INDICADOR", "DATA ULT AT"])
#Caminho absoluto
absPath = str(Path(__file__).parent.absolute()) + "\\"
#==================================================#

#============== CONFIGURACOES DRIVER ==============#
#Iniciando o navegador
driver_path = absPath + "geckodriver.exe" 
profile_path = r"C:\Users\<>\AppData\Roaming\Mozilla\Firefox\Profiles\<>"
options = Options()
options.headless = True #False 
options.add_argument("-profile") 
options.add_argument(profile_path)  
service = Service(driver_path)
driver = Firefox(service=service, options=options)
wait = WebDriverWait(driver, 800) 
driver.get("http://web.whatsapp.com")
#==================================================#

#========== CONFIGURAÇÕES PLANILHAS ===============#
lstPlanilhas = ["Exemplo.xlsx"]
lstAbasPlanilhas = ["Aba1"]
cordCont = [1, 2]   
cordFig = [2, 2]    
cordDts = [4, 2]   
lstContsAdm = ["<>"]
#==================================================#

def main():
    try:
        for plan, aba in zip(lstPlanilhas, lstAbasPlanilhas):
            rotina(plan, aba, cordCont, cordFig, cordDts, lstContsAdm)
    except Exception as e:
        print(f"Erro na função main: {e}")
    #Finalizando o driver
    finally:
        time.sleep(5)
        driver.close()


def rotina(nomePlan, nomeAbaPlan, cordContatos, cordFiguras, cordDatas, lstContatosAdm):
    print(f"\nRotina ### {nomePlan} ###\n")
    try:
        atualiza_planilha(nomePlan)
        importar_dados(nomePlan, nomeAbaPlan, cordContatos, cordFiguras, cordDatas)
        print(lstContatos)
        print(lstFiguras)
        salvar_imagem(nomePlan, nomeAbaPlan)
        envia_imagens(lstContatosAdm)
    except Exception as e:
        print(f"Erro na rotina: {e}")
        envia_mensagem(lstContatosAdm, f"Erro na execução da rotina da planilha *{nomePlan}*")
        envia_mensagem(lstContatosAdm, f"Descrição do erro:\n{e}")

    numRows = 0

    for row in tabAtrasos:
        numRows+=1

    if numRows>0:
        print("Enviando tabela indicadores desatualizados...")
        envia_mensagem(
            lstContatosAdm, f"*{nomePlan}* " + str(datetime.today().strftime('%H:%M - %d-%m-%Y')) + ".\n"
            "Indicadores Desatualizados:\n" + 
            "```" + str(tabAtrasos) + "```"
        )
    else:
        envia_mensagem(
            lstContatosAdm, f"*{nomePlan}* {datetime.today().strftime('%H:%M - %d-%m-%Y')}.\n" + 
            f"{contEnvios-len(lstContatosErros)} Indicadores Enviados " 
        )

    if len(lstContatosErros)>0:
        envia_mensagem(
            lstContatosAdm, f"Contatos c/ Erros:\n" +
            str(lstContatosErros)
        )
    deleta_arquivos()


def atualiza_planilha(nomePlan):
    print(f"Atualizando dados '{nomePlan}'...")

    excel = win32.DispatchEx("Excel.Application")
    wb = excel.Workbooks.Open(absPath + nomePlan)
    excel.DisplayAlerts = False
    try:
        wb.RefreshAll()
    except Exception as e:
        print(f"Erro na funcao 'atualiza_planilha'\n{e}")
    finally:
        wb.Save()
        excel.Quit()


def importar_dados(nomePlan, nomeAbaPlan, cordContatos, cordFiguras, cordDatas):
    print("Importando dados...")
    wb = load_workbook(nomePlan, read_only=True, keep_vba=True)
    ws = wb[nomeAbaPlan] 

    try:
        for row in range(cordContatos[Linha], ws.max_row+1):
            if(ws.cell(row, cordContatos[Coluna]).value is None):
                break
            elif(ws.cell(row, cordFiguras[Coluna]).value is None):
                break
            elif(ws.cell(row, cordDatas[Coluna]).value is None):
                break
            else:
                lstContatos.append(ws.cell(row, cordContatos[Coluna]).value) 
                lstFiguras.append(ws.cell(row, cordFiguras[Coluna]).value)
                lstDatas.append(ws.cell(row, cordDatas[Coluna]).value.strftime(r"%d/%m/%Y %H:%M"))
    except Exception as e:
        print(f"Erro na funcao 'importar_dados': {e}")
    finally:
        wb.close()
        

def salvar_imagem(nomePlan, nomeAbaPlan):
    print("Salvando imagens...")
    excel = win32.DispatchEx("Excel.Application")
    wb = excel.Workbooks.Open(absPath + nomePlan, ReadOnly=True)
    excel.DisplayAlerts = False

    try:
        MYDIR = str(Path(__file__).parent.absolute()) + "\\temp"
        CHECK_FOLDER = os.path.isdir(MYDIR)

        if not CHECK_FOLDER:
            os.makedirs(MYDIR)
            print("Pasta 'temp' criada.")

        sheet = excel.Sheets(nomeAbaPlan)

        for i, shape in enumerate(sheet.Shapes): 
            for figura in lstFiguras: 
                if shape.Name.startswith(figura): 
                    shape.Copy()
                    image = ImageGrab.grabclipboard() 
                    image = image.convert('RGB') 
                    image.save(absPath + "\\temp\\" + figura + ".jpg", "jpeg")
    except Exception as e:
        print(f"Erro na funcao 'salvar_imagem': {e}")
    finally:
        wb.Close(False)
        excel.Quit()


def envia_imagens(lstContatosAdm):
    print("Enviando mensagens...")
    global contEnvios
    for i in range(0, len(lstContatos)):
        try:
            if datetime.strptime(lstDatas[i], r"%d/%m/%Y %H:%M") < datetime.today() - timedelta(days=1):
                tabAtrasos.add_row([lstFiguras[i], lstDatas[i]])
        
            x_arg = '//span[contains(@title, ' + '"' + str(lstContatos[i]) + '"' + ')]'
            group_title = wait.until(EC.presence_of_element_located((
                By.XPATH, x_arg)))
            group_title.click()

            driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
            attach = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
            try:
                attach.send_keys(absPath + "\\temp\\" + str(lstFiguras[i]) + ".jpg")
                time.sleep(1.5)
            except:
                print(f"Erro na figura {lstFiguras[i]}.")
                input_path = '//div[@contenteditable="true"][@data-tab="10"]'
                input_box = wait.until(EC.presence_of_element_located((
                    By.XPATH, input_path)))
                input_box.send_keys(Keys.ESCAPE)
                next
                
            send = driver.find_element(By.CSS_SELECTOR, "span[data-icon='send']")
            time.sleep(1.5)

            send.click()

            time.sleep(1.5)

            print("Enviando imagem " + str(i+1) + ": ## " + str(lstFiguras[i]) + " ## Para: ## " + str(lstContatos[i]) + " ##")
            contEnvios += 1
        except Exception as e:
            lstContatosErros.append(str(lstContatos[i] + " - " + lstFiguras[i]))
            print(f"Erro no modulo 'envia_imagens' p/ {lstContatos[i]}: {e}")
            """
            envia_mensagem(lstContatosAdm, f"Erro no modulo 'envia_imagens'.\n" +
            f"Contato -> {lstContatos[i]}\n" +
            f"Descrição do errro:{e}")
            """


def envia_mensagem(lstContatoMsg, mensagem):
    for i in range(0, len(lstContatoMsg)):
        try:
            print(f"Enviando mensagem p/ {lstContatoMsg[i]}...")        
            x_arg = '//span[contains(@title, ' + '"' + str(lstContatoMsg[i]) + '"' + ')]'
            group_title = wait.until(EC.presence_of_element_located((
                By.XPATH, x_arg)))
            group_title.click()
            time.sleep(1.5)
            input_path = '//div[@contenteditable="true"][@data-tab="10"]'
            input_box = wait.until(EC.presence_of_element_located((
                By.XPATH, input_path)))
            time.sleep(1.5)
            for line in mensagem.split('\n'):
                input_box.send_keys(line)
                input_box.send_keys(Keys.SHIFT, Keys.ENTER)

            time.sleep(1.5)
            input_box.send_keys(Keys.ENTER)

            print(f"Mensagem enviada para -> {lstContatoMsg[i]}")
        except Exception as e:
            print(f"Erro no modulo 'envia_mensagens' p/ {lstContatoMsg[i]}: {e}")


def deleta_arquivos():
    lstContatos.clear()
    lstFiguras.clear()
    lstDatas.clear()
    lstContatosErros.clear()
    global contEnvios
    try: 
        shutil.rmtree(absPath + "\\temp")
        print("Pasta 'temp' encontrada e deletada.")
    except:
        print("Pasta 'temp' não encontrada.")

    try:
        shutil.rmtree(os.path.expandvars(r'%TEMP%\gen_py'))
        print("Pasta 'gen_py' encontrada e deletada.")
    except:
        print("Pasta 'gen_py' não encontrada.")

    try:
        tabAtrasos.clear_rows()
        print("Tabela atrasos limpa.")
    except:
        print("Erro ao apagar linhas da PrettyTable.")

    contEnvios = 0


if __name__=="__main__":
    main()