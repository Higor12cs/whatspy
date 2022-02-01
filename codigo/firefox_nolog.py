import os, time, shutil
from pathlib import Path

from openpyxl import load_workbook
from PIL import ImageGrab
import win32com.client as win32

from selenium.webdriver import Firefox
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.keys import Keys 

from datetime import datetime, timedelta
from prettytable import PrettyTable

def importar_dados():
    print("Importando dados 'importar_dados'...")

    wb = load_workbook(planilha_contatos, read_only=True, keep_vba=True)
    ws = wb[aba_planilha] 
    
    for row in range(linha_contatos, ws.max_row+1):
        if(ws.cell(row, coluna_contatos).value is None):
            break
        elif(ws.cell(row, coluna_figuras).value is None):
            break
        elif(ws.cell(row, coluna_datas).value is None):
            break
        else:
            lista_contatos.append(ws.cell(row, coluna_contatos).value) 
            lista_figuras.append(ws.cell(row, coluna_figuras).value)
            lista_datas.append(ws.cell(row, coluna_datas).value.strftime(r"%d/%m/%Y %H:%M"))

    wb.close()
        

def salvar_imagem():
    print("Salvando imagens...")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(planilha_contatos)

    MYDIR = str(Path(__file__).parent.absolute()) + "\\temp"
    CHECK_FOLDER = os.path.isdir(MYDIR)

    if not CHECK_FOLDER:
        os.makedirs(MYDIR)
        print("Pasta 'temp' criada.")

    sheet = excel.Sheets(aba_planilha)

    for i, shape in enumerate(sheet.Shapes): 
        for figura in lista_figuras: 
            if shape.Name.startswith(figura): 
                shape.Copy()
                image = ImageGrab.grabclipboard() 
                image = image.convert('RGB') 
                image.save(str(Path(__file__).parent.absolute()) + "\\temp\\" + figura + ".jpg", "jpeg")

    #excel.DisplayAlerts = False
    wb.Close(False)
    excel.Quit()

def envia_imagens():
    print("Enviando mensagens...")
    
    for i in range(0, len(lista_contatos)):
        if datetime.strptime(lista_datas[i], r"%d/%m/%Y %H:%M") < datetime.today() - timedelta(days=1):
            prettytable.add_row([lista_figuras[i], lista_datas[i]])

        print("Enviando imagem " + str(i+1) + ": ## " + str(lista_figuras[i]) + " ## Para: ## " + str(lista_contatos[i]) + " ##")
    
        x_arg = '//span[contains(@title, ' + '"' + str(lista_contatos[i]) + '"' + ')]'
        group_title = wait.until(EC.presence_of_element_located((
            By.XPATH, x_arg)))
        group_title.click()
        driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
        time.sleep(1)
        attach = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
        attach.send_keys(str(Path(__file__).parent.absolute()) + "\\temp\\" + str(lista_figuras[i]) + ".jpg")
        time.sleep(1)
        send = driver.find_element(By.CSS_SELECTOR, "span[data-icon='send']")
        send.click()
        time.sleep(1)

def envia_tabela_atrasos():
    print("Enviando resultados...")

    for i in range(0, len(contatos_adm)):        
        x_arg = '//span[contains(@title, ' + '"' + str(contatos_adm[i]) + '"' + ')]'
        group_title = wait.until(EC.presence_of_element_located((
            By.XPATH, x_arg)))
        group_title.click()

        input_path = '//div[@contenteditable="true"][@data-tab="10"]'
        input_box = wait.until(EC.presence_of_element_located((
            By.XPATH, input_path)))
        
        mensagem = "*Indicadores desatualizados:*\n" + "```" + str(prettytable) + "```"

        for line in mensagem.split('\n'):
            input_box.send_keys(line)
            input_box.send_keys(Keys.SHIFT, Keys.ENTER)

        input_box.send_keys(Keys.ENTER)
        print("Tabela enviada para -> " + contatos_adm[i])

def envia_mensagem(mensagem):
    for i in range(0, len(contatos_adm)):        
        x_arg = '//span[contains(@title, ' + '"' + str(contatos_adm[i]) + '"' + ')]'
        group_title = wait.until(EC.presence_of_element_located((
            By.XPATH, x_arg)))
        group_title.click()

        input_path = '//div[@contenteditable="true"][@data-tab="10"]'
        input_box = wait.until(EC.presence_of_element_located((
            By.XPATH, input_path)))
        
        for line in mensagem.split('\n'):
            input_box.send_keys(line)
            input_box.send_keys(Keys.SHIFT, Keys.ENTER)

        input_box.send_keys(Keys.ENTER)

        print("Mensagem enviada para -> " + contatos_adm[i])

def deleta_arquivos():
    try: 
        shutil.rmtree(str(Path(__file__).parent.absolute()) + "\\temp")
        print("Pasta 'temp' encontrada e deletada.")
    except Exception as e:
        print("Pasta 'temp' não encontrada: " + str(e))

    try:
        shutil.rmtree(os.path.expandvars(r'%TEMP%\gen_py'))
        print("Pasta 'gen_py' encontrada e deletada.")
    except Exception as e:
        print("Pasta 'gen_py' não encontrada: " + str(e))

def rotina():
    try:
        importar_dados()
        salvar_imagem() 
        time.sleep(5)
        envia_imagens() 
        time.sleep(5)

        aux = 0
        for row in prettytable:
            aux+=1

        if aux == 0:
            envia_mensagem("*Indicadores_00:00*\n\n" + str(len(lista_contatos)) + " indicadores enviados às " + str(datetime.today().strftime('%d-%m-%Y %H:%M')) +".")
        else:
            envia_mensagem("*Indicadores_00:00*\n\n" + str(len(lista_contatos)) + " indicadores enviados às " + str(datetime.today().strftime('%d-%m-%Y %H:%M')) +".")
            envia_mensagem(str(lista_contatos))
            envia_mensagem("*Atenção:* " + str(aux) + " indicador(es) com atraso(s).")
            envia_tabela_atrasos()

    except Exception as e:
        print("Erro executando a rotina, enviando mensagem de erro para os contatos na lista 'contatos_adm'.")
        mensagem = "Erro nos Indicadores_00:00:\n" + "```" + str(e) + "```"
        envia_mensagem(mensagem)
        time.sleep(5)
    finally:
        deleta_arquivos()

lista_contatos = []
lista_figuras = []
lista_datas = []

contatos_adm = ["<>"]
driver_path = str(Path(__file__).parent.absolute()) + r"\geckodriver.exe" 
profile_path = r"C:\Users\<>\AppData\Roaming\Mozilla\Firefox\Profiles\<>"
planilha_contatos = str(Path(__file__).parent.absolute()) + r"\Exemplo.xlsx"   
aba_planilha = "Aba1"
coluna_contatos = 1     
linha_contatos = 2 
coluna_figuras = 2 
linha_figuras = 2
coluna_datas = 3
linha_datas = 2

prettytable = PrettyTable(["INDICADOR", "DATA ULT AT"])

print("Iniciando Driver")
options = Options()
options.headless = True 
options.add_argument("-profile") 
options.add_argument(profile_path)  

service = Service(driver_path)

driver = Firefox(service=service, options=options) 

driver.get("http://web.whatsapp.com")

wait = WebDriverWait(driver, 800)

if os.path.exists(planilha_contatos):
    print("Planilha = True")
    rotina()
else:
    print("Arquivo Excel nao existe!")
    envia_mensagem("Planilha *'Indicadores_00:00'* não encontrada!")

time.sleep(5)

driver.close()
