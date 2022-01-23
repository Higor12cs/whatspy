import os, time, pyperclip, logging, shutil
from pathlib import Path

from openpyxl import load_workbook
from PIL import ImageGrab
import win32com.client as win32

from selenium.webdriver import Firefox
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.keys import Keys 

from datetime import datetime, timedelta
from prettytable import PrettyTable

start_time = datetime.now() #iniciando timer do script

def importar_dados(planilha):
    print("Importando dados 'importar_dados'...")
    logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Importando dados 'importar_dados'...")
    
    try:
        wb = load_workbook(planilha, read_only=True, keep_vba=True)
        ws = wb[aba_planilha] #nome da aba da planilha
        
        for row in range(linha_contatos, ws.max_row+1): #numero da linha onde começa / maximo de linhas +1 para não finalizar antes do ultimo elemento   
            #contatos
            if(ws.cell(row, coluna_contatos).value is None): #se valor é nulo -> break
                break
            else:
                lista_contatos.append(ws.cell(row, coluna_contatos).value) 
            #figuras
            if(ws.cell(row, coluna_figuras).value is None): #se valor é nulo -> break
                break
            else:
                lista_figuras.append(ws.cell(row, coluna_figuras).value)
            #datas
            if(ws.cell(row, coluna_datas).value is None): #se valor é nulo -> break
                break
            else:
                lista_datas.append(ws.cell(row, coluna_datas).value.strftime(r"%d/%m/%Y %H:%M"))
    except Exception as e:
        print("Erro no modulo 'importar_dados': " + str(e))
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Erro no modulo 'importar_dados': " + str(e))
    finally:
        wb.close()
        print("Lista de contatos importada:") 
        print(lista_contatos)
        print("### " + str(len(lista_contatos)) + " Contatos importados ###")
        logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Lista de contatos importada:") 
        logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === ### " + str(len(lista_contatos)) + " Contatos importados ###") 
        logging.info(lista_contatos)

        print("Lista de figuras importadas:")
        print(lista_figuras)
        print("### " + str(len(lista_figuras)) + " Figuras importadas ###")
        logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Lista de figuras importadas:")
        logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === ### " + str(len(lista_figuras)) + " Figuras importadas ###")
        logging.info(lista_figuras)

        print("Lista de datas importadas:")
        print(lista_datas)
        print("### " + str(len(lista_datas)) + " Datas importadas ###")
        logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Lista de datas importadas:")
        logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === ### " + str(len(lista_datas)) + " Datas importadas ###")
        logging.info(lista_datas)

def salvar_imagem(planilha):
    print("Salvando imagens...")
    logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Salvando imagens...")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(planilha)

    try:
        MYDIR = str(Path(__file__).parent.absolute()) + "\\temp"
        CHECK_FOLDER = os.path.isdir(MYDIR)

        if not CHECK_FOLDER:
            os.makedirs(MYDIR)
            print("Pasta 'temp' criada.")
            logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Pasta 'temp' criada.")

        sheet = excel.Sheets(aba_planilha)

        for i, shape in enumerate(sheet.Shapes): #percorre todos os shapes da pasta selecionada
            for figura in lista_figuras: #percorre todas as figuras armazenadas na lista de figuras
                if shape.Name.startswith(figura): #se o nome da figura for igual ao elemento da lista de figuras
                    shape.Copy()
                    image = ImageGrab.grabclipboard() #copia
                    image = image.convert('RGB') #converte de RGBA p/ RGB
                    image.save(str(Path(__file__).parent.absolute()) + "\\temp\\" + figura + ".jpg", "jpeg") #salva com a extensao .jpg
    except Exception as e:
        print("Erro no modulo 'salvar_imagem': " + str(e))
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Erro no modulo 'salvar_imagem': " + str(e))
    finally:
        #excel.DisplayAlerts = False
        wb.Close(False) #fecha o workbook
        excel.Quit()

def envia_imagens(contatos, imagens, datas):
    print("Enviando mensagens...")
    logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Enviando mensagens...")
    
    try:
        for i in range(0, len(contatos)): #percorre do 0 até o tamanho do array dos contatos enviando para o contato a respectiva imagem na mesma linha da planilha
            #verificando data de atualizacao dos indicadores
            if datetime.strptime(lista_datas[i], r"%d/%m/%Y %H:%M") < datetime.today() - timedelta(days=1):
                table.add_row([imagens[i], datas[i]])
                #print(lista_datas[i])
                #print(datetime.strptime(lista_datas[i], r"%d/%m/%Y %H:%M"))

            print("Enviando imagem " + str(i+1) + ": ## " + str(imagens[i]) + " ## Para: ## " + str(contatos[i]) + " ##")
            logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Enviando imagem " + str(i+1) + ": ## " + str(imagens[i]) + " ## Para: ## " + str(contatos[i]) + " ##")
            #procura na pagina o elemento com o mesmo nome do contato
            x_arg = '//span[contains(@title, ' + '"' + str(contatos[i]) + '"' + ')]'
            group_title = wait.until(EC.presence_of_element_located((
                By.XPATH, x_arg)))
            group_title.click()
            #procura na pagina o elemento de anexo de arquivos
            driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
            time.sleep(1)
            attach = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
            #passa o caminho da imagem para ser anexada
            attach.send_keys(str(Path(__file__).parent.absolute()) + "\\temp\\" + str(imagens[i]) + ".jpg")
            time.sleep(1)
            #tecla enviar
            send = driver.find_element(By.CSS_SELECTOR, "span[data-icon='send']")
            send.click()
            time.sleep(1)
    except Exception as e:
        print("Erro no modulo 'envia_imagens': " + str(e))
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Erro no modulo 'envia_imagens': " + str(e))

def envia_tabelaAtrasos(ptablle, contatos):
    print("Enviando resultados...")
    try:
        for i in range(0, len(contatos)):
            logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Enviando resultados...")
            
            x_arg = '//span[contains(@title, ' + '"' + str(contatos[i]) + '"' + ')]'
            group_title = wait.until(EC.presence_of_element_located((
                By.XPATH, x_arg)))
            group_title.click()

            #procura na pagina o elemento de anexo de arquivos
            input_path = '//div[@contenteditable="true"][@data-tab="10"]'
            input_box = wait.until(EC.presence_of_element_located((
                By.XPATH, input_path)))
            
            mensagem = "Indicadores desatualizados:\n" + "```" + str(ptablle) + "```"

            for line in mensagem.split('\n'):
                input_box.send_keys(line)
                input_box.send_keys(Keys.SHIFT, Keys.ENTER)

            #pyperclip.copy(mensagem)
            input_box.send_keys(Keys.ENTER)
            #pyperclip.copy("")
            print("Tabela enviada para -> " + contatos[i])
            logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Tabela enviada para -> " + contatos[i])
    except Exception as e:
        print("Erro no modulo 'envia_tabelaAtrasos': " + str(e))
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Erro no modulo 'envia_tabelaAtrasos': " + str(e))

def deleta_arquivos():
    #deletando pasta temp de imagens
    try: 
        shutil.rmtree(str(Path(__file__).parent.absolute()) + "\\temp")
        print("Pasta 'temp' encontrada e deletada.")
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Pasta 'temp' encontrada e deletada.")
    except Exception as e:
        print("Pasta 'temp' não encontrada: " + str(e))
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Pasta 'temp' não encontrada: " + str(e))

    #deletando pasta %temp%/gen_py
    try: 
        shutil.rmtree(os.path.expandvars(r'%TEMP%\gen_py'))
        print("Pasta 'gen_py' encontrada e deletada.")
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Pasta 'gen_py' encontrada e deletada.")
    except Exception as e:
        print("Pasta 'gen_py' não encontrada: " + str(e))
        logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Pasta 'gen_py' não encontrada: " + str(e))

def rotina():
    #=== EXECUTANDO FUNÇÕES ===#
    importar_dados(planilha_contatos)
    salvar_imagem(planilha_contatos) 
    time.sleep(5)
    envia_imagens(lista_contatos, lista_figuras, lista_datas) 
    time.sleep(5)
    envia_tabelaAtrasos(table, contatos_adm)
    deleta_arquivos()

#=== INICIANDO LOG ===#
logging.basicConfig(filename=str(Path(__file__).parent.absolute()) + "\\" + "historico.log", level=logging.INFO, format='%(message)s')
logging.info("")
logging.info("*** INICIO - " + str(datetime.now().strftime(r'%Y-%m-%d %H:%M:%S')))

#=== CRIANDO UM PROFILE NO FIREFOX ===#

# Windows + R
# "firefox.exe -p"
# Crie um novo perfil
# Entre na pasta: "%appdata%\Roaming\Mozilla\Firefox\Profiles"
# Copie o caminho para a variavel "profile_path"

lista_contatos = []
lista_figuras = []
lista_datas = []

contatos_adm = ["Contato"]
driver_path = str(Path(__file__).parent.absolute()) + r"\geckodriver.exe" 
profile_path = r"C:\Users\<seu_usuario>\AppData\Roaming\Mozilla\Firefox\Profiles\<seu_profile>"
planilha_contatos = str(Path(__file__).parent.absolute()) + r"\Exemplo.xlsx"   
aba_planilha = "Aba1"
coluna_contatos = 1     
linha_contatos = 2 
coluna_figuras = 2 
linha_figuras = 2
coluna_datas = 3
linha_datas = 2

#=== INICIANDO A TABELA DE RESULTADOS ===#
table = PrettyTable(["INDICADOR", "DATA ULT AT"])

#=== INICIANDO O NAVEGADOR ===#
options = Options()
options.headless = True #true pra rodar o firefox em modo oculto
options.add_argument("-profile") 
options.add_argument(profile_path) #adicionando a pasta profile p/ o whatsapp não ficar pedindo o QR code  

service = Service(driver_path) #caminho do driver

driver = Firefox(service=service, options=options) 

driver.get("http://web.whatsapp.com")
logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Abrindo driver no WhatsApp")
print("Abrindo driver no WhatsApp")

wait = WebDriverWait(driver, 800)
logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Wait...")  

#=== VERIFICANDO SE EXISTE O ARQUIVO EXCEL ===#

if os.path.exists(planilha_contatos):
    print("Planilha = True")
    logging.info(str(datetime.now().strftime(r'%H:%M:%S')) + " === Planilha = True")
    rotina()
else:
    print("Arquivo Excel nao existe!")
    logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Arquivo Excel nao existe! Verifique se '" + planilha_contatos + "' esta correto.")
    driver.close() #fecha o driver

end_time = datetime.now()

time.sleep(5)
try:    
    driver.close()
except Exception as e:
    print("Erro ao fechar o driver: " + str(e))
    logging.error(str(datetime.now().strftime(r'%H:%M:%S')) + " === Erro ao fechar o driver: " + str(e))

#logging final e tempo de execucao
logging.info("*** Tempo de execução: {}".format(end_time - start_time))
print('*** Tempo de execução: {}'.format(end_time - start_time))

logging.info("*** FIM *** - " + str(datetime.now().strftime(r'%Y-%m-%d %H:%M:%S')))